"""
report_room.py
==============
Export the room-shielding result as a report the user picks the format for:
PDF (default), Excel, or HTML. All three are built from one `build_report()` dict
so the content is identical across formats.

PDF uses fpdf2 (pure-python — no system libraries, so it deploys on Streamlit
Cloud). Excel uses openpyxl with a live margin formula. HTML is a self-contained
string with the diagram inlined as base64.
"""

from __future__ import annotations

import base64
import datetime as _dt
import io
from typing import Dict, List

from .model import RoomDesign, WALL_NAMES
from .engines import EngineResult
from .decision_support import explain_failures, summarize_results

DISCLAIMER = (
    "Decision-support output — requires review and sign-off by a qualified Radiation "
    "Safety Officer / medical physicist. Analytical tier: NCRP-151 / AAPM TG-108 "
    "broad-beam model (ShieldLab, validated). Surrogate tier (where shown): "
    "Monte-Carlo-trained Extra-Trees with conformal 95% intervals and an "
    "out-of-domain guard (thesis, 2026)."
)


def _fmt(x, nd=3):
    if x is None:
        return "—"
    if isinstance(x, bool):
        return "PASS" if x else "FAIL"
    if isinstance(x, float):
        if x == 0:
            return "0"
        if abs(x) < 1e-3 or abs(x) >= 1e4:
            return f"{x:.2e}"
        return f"{x:.{nd}g}"
    return str(x)


def build_report(design: RoomDesign, results: List[EngineResult],
                 mode: str, diagram_png: bytes,
                 surrogate_results: List[EngineResult] = None) -> Dict:
    s = design.source
    smap = {r.label: r for r in (surrogate_results or [])}
    rows = []
    for r in results:
        sr = smap.get(r.label)
        prim = sr if sr is not None else r        # verdict source = geometry-aware tier
        if sr is not None and sr.ci_low is not None:
            s_b = _fmt(sr.B_achieved)
            s_ci = f"[{sr.ci_low:.1e}, {sr.ci_high:.1e}]"
            tier = sr.engine
        elif sr is not None:
            s_b, s_ci, tier = _fmt(sr.B_achieved), "—", sr.engine
        else:
            s_b, s_ci, tier = "—", "—", r.engine
        rows.append({
            "barrier": r.label,
            "material": (prim.material or r.material) or "—",
            "suggested_mm": _fmt(r.suggested_thickness_mm, 4) if mode == "design" else "—",
            "B_required": _fmt(r.B_required),
            "B_achieved": _fmt(r.B_achieved),
            "B_surrogate": s_b,
            "surrogate_CI95": s_ci,
            "tier": tier,
            "dose_mSv_wk": _fmt(prim.dose_mSv_wk),
            "limit_mSv_wk": _fmt(prim.goal_over_T),
            "verdict": _fmt(prim.passes),
            "margin": _fmt(prim.margin, 3),
            "engine": prim.engine,
            "note": prim.note,
        })
    primary_results = [smap.get(result.label, result) for result in results]
    return {
        "title": "ShieldLab — Room Shielding Report",
        "timestamp": _dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "mode": "Design (suggested shielding)" if mode == "design" else "Check (declared shielding)",
        "inputs": {
            "Isotope": s.isotope,
            "Activity per patient (MBq)": s.activity_MBq,
            "Patients / week": s.patients_per_week,
            "Residence (min/patient)": s.residence_min,
            "Room (W×L×H, m)": f"{design.room.width_m:g} × {design.room.length_m:g} × {design.room.height_m:g}",
            "Source position (x,y m)": f"({s.x_m:g}, {s.y_m:g})",
            "Framework": design.framework,
        },
        "rows": rows,
        "summary": summarize_results(primary_results),
        "failure_explanations": explain_failures(design, primary_results),
        "diagram_png": diagram_png,
        "disclaimer": DISCLAIMER,
        "notes": design.notes,
    }


# --------------------------------------------------------------------------- PDF
def _register_unicode_font(pdf) -> str:
    """Register the DejaVu Unicode font (bundled with matplotlib) so the PDF can
    render em-dashes, ×, ≤ etc. Returns the family name, or 'Helvetica' if the
    font files cannot be found (falling back to ASCII-sanitised text)."""
    import os
    import matplotlib
    ttf = os.path.join(matplotlib.get_data_path(), "fonts", "ttf")
    reg = os.path.join(ttf, "DejaVuSans.ttf")
    bold = os.path.join(ttf, "DejaVuSans-Bold.ttf")
    obl = os.path.join(ttf, "DejaVuSans-Oblique.ttf")
    if not os.path.exists(reg):
        return "Helvetica"
    pdf.add_font("DejaVu", "", reg)
    pdf.add_font("DejaVu", "B", bold if os.path.exists(bold) else reg)
    pdf.add_font("DejaVu", "I", obl if os.path.exists(obl) else reg)
    return "DejaVu"


def to_pdf(report: Dict) -> bytes:
    from fpdf import FPDF

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    FONT = _register_unicode_font(pdf)

    def sane(t):  # only needed if the Unicode font was unavailable
        return t if FONT == "DejaVu" else (
            str(t).replace("—", "-").replace("×", "x").replace("≤", "<=")
            .replace("·", "-").encode("latin-1", "replace").decode("latin-1"))

    pdf.set_font(FONT, "B", 15)
    report = {**report, "_FONT": FONT}
    pdf.cell(0, 9, sane(report["title"]), ln=1)
    pdf.set_font(FONT, "", 9)
    pdf.set_text_color(90, 90, 90)
    pdf.cell(0, 5, sane(f"{report['mode']}    |    generated {report['timestamp']}"), ln=1)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    # inputs
    pdf.set_font(FONT, "B", 11)
    pdf.cell(0, 6, "Inputs", ln=1)
    pdf.set_font(FONT, "", 9)
    for k, v in report["inputs"].items():
        pdf.cell(70, 5, sane(f"{k}:"), border=0)
        pdf.cell(0, 5, sane(str(v)), ln=1)
    pdf.ln(1)

    # diagram
    if report.get("diagram_png"):
        img = io.BytesIO(report["diagram_png"])
        try:
            pdf.image(img, w=110)
        except Exception:
            pass
    pdf.ln(2)

    # results table
    pdf.set_font(FONT, "B", 11)
    pdf.cell(0, 6, "Per-barrier results", ln=1)
    headers = [("Barrier", 28), ("Mat.", 16), ("Sug. mm", 16), ("Analyt. B", 20),
               ("Surrogate B (95% CI)", 44), ("Dose", 20), ("Verdict", 16), ("Margin", 14)]
    pdf.set_font(FONT, "B", 8)
    pdf.set_fill_color(230, 234, 240)
    for h, w in headers:
        pdf.cell(w, 6, sane(h), border=1, fill=True, align="C")
    pdf.ln(6)
    pdf.set_font(FONT, "", 8)
    for row in report["rows"]:
        v = row["verdict"]
        if v == "PASS":
            pdf.set_text_color(30, 110, 45)
        elif v == "FAIL":
            pdf.set_text_color(190, 30, 30)
        else:
            pdf.set_text_color(120, 120, 120)
        s_col = row["B_surrogate"]
        if row["surrogate_CI95"] != "—":
            s_col = f"{row['B_surrogate']} {row['surrogate_CI95']}"
        cells = [(row["barrier"], 28), (row["material"], 16), (row["suggested_mm"], 16),
                 (row["B_achieved"], 20), (s_col, 44), (row["dose_mSv_wk"], 20),
                 (v, 16), (row["margin"], 14)]
        for txt, w in cells:
            pdf.cell(w, 6, sane(str(txt)[:30]), border=1, align="C")
        pdf.ln(6)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    # disclaimer
    pdf.set_font(FONT, "I", 7.5)
    pdf.set_text_color(90, 90, 90)
    pdf.multi_cell(0, 4, sane(report["disclaimer"]))
    if report.get("notes"):
        pdf.ln(1)
        pdf.multi_cell(0, 4, sane(f"Notes: {report['notes']}"))

    out = pdf.output()
    return bytes(out)


def to_summary_pdf(report: Dict) -> bytes:
    """Create a compact, single-page clinical decision summary."""
    from fpdf import FPDF

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()
    font = _register_unicode_font(pdf)

    def sane(text):
        return text if font == "DejaVu" else str(text).encode("latin-1", "replace").decode("latin-1")

    summary = report["summary"]
    critical = summary["critical"]
    colors = {"PASS": (31, 122, 61), "FAIL": (190, 45, 45), "MARGINAL": (185, 122, 0)}
    red, green, blue = colors[summary["status"]]

    pdf.set_font(font, "B", 16)
    pdf.cell(0, 8, sane("ShieldLab | Room Shielding Decision Summary"), ln=1)
    pdf.set_font(font, "", 8.5)
    pdf.set_text_color(85, 85, 85)
    pdf.cell(0, 5, sane(f"{report['mode']} | generated {report['timestamp']}"), ln=1)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    pdf.set_fill_color(red, green, blue)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(font, "B", 18)
    pdf.cell(0, 14, summary["status"], fill=True, align="C", ln=1)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font(font, "", 9)
    pdf.multi_cell(0, 5, sane(summary["message"]), align="C")
    pdf.ln(2)

    pdf.set_font(font, "B", 11)
    pdf.cell(0, 6, "Critical barrier result", ln=1)
    pdf.set_font(font, "", 9)
    if critical:
        pdf.set_fill_color(245, 247, 250)
        pdf.cell(55, 7, "Barrier", border=1, fill=True)
        pdf.cell(0, 7, sane(critical.label), border=1, ln=1)
        pdf.cell(55, 7, "Calculated dose", border=1, fill=True)
        pdf.cell(0, 7, sane(f"{critical.dose_mSv_wk:.3g} mSv/week"), border=1, ln=1)
        pdf.cell(55, 7, "Regulatory design goal", border=1, fill=True)
        pdf.cell(0, 7, sane(f"{critical.goal_over_T:.3g} mSv/week"), border=1, ln=1)
        pdf.cell(55, 7, "Safety margin", border=1, fill=True)
        pdf.cell(0, 7, sane(f"{critical.margin:.2f}x" if critical.margin is not None else "Not available"),
                 border=1, ln=1)
        if critical.ci_low is not None and critical.ci_high is not None:
            low_dose = critical.dose_mSv_wk * critical.ci_low / critical.B_achieved
            high_dose = critical.dose_mSv_wk * critical.ci_high / critical.B_achieved
            pdf.cell(55, 7, "AI 95% confidence interval", border=1, fill=True)
            pdf.cell(0, 7, sane(f"{low_dose:.3g} to {high_dose:.3g} mSv/week"), border=1, ln=1)
    else:
        pdf.multi_cell(0, 5, "No barrier path could be evaluated.")

    pdf.ln(3)
    pdf.set_font(font, "B", 11)
    pdf.cell(0, 6, "Room inputs", ln=1)
    pdf.set_font(font, "", 8.5)
    inputs = list(report["inputs"].items())
    for index in range(0, len(inputs), 2):
        pair = inputs[index:index + 2]
        for key, value in pair:
            pdf.set_font(font, "B", 8.5)
            pdf.cell(45, 5, sane(f"{key}:"))
            pdf.set_font(font, "", 8.5)
            pdf.cell(50, 5, sane(str(value)))
        if len(pair) == 1:
            pdf.cell(95, 5, "")
        pdf.ln(5)

    pdf.ln(3)
    pdf.set_font(font, "B", 11)
    pdf.cell(0, 6, "Barrier status", ln=1)
    headers = [("Barrier", 54), ("Dose (mSv/wk)", 36), ("Goal", 32), ("Verdict", 28), ("Tier", 40)]
    pdf.set_font(font, "B", 7.5)
    pdf.set_fill_color(48, 84, 150)
    pdf.set_text_color(255, 255, 255)
    for header, width in headers:
        pdf.cell(width, 5.5, header, border=1, fill=True, align="C")
    pdf.ln(5.5)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font(font, "", 7.5)
    for row in report["rows"][:8]:
        pdf.cell(54, 5.5, sane(row["barrier"][:28]), border=1)
        pdf.cell(36, 5.5, sane(row["dose_mSv_wk"]), border=1, align="C")
        pdf.cell(32, 5.5, sane(row["limit_mSv_wk"]), border=1, align="C")
        pdf.cell(28, 5.5, sane(row["verdict"]), border=1, align="C")
        pdf.cell(40, 5.5, sane(row["tier"][:22]), border=1, align="C")
        pdf.ln(5.5)

    if report["failure_explanations"]:
        pdf.ln(2)
        pdf.set_font(font, "B", 10)
        pdf.cell(0, 5, "Failure explanation", ln=1)
        pdf.set_font(font, "", 8)
        explanation = report["failure_explanations"][0]
        pdf.multi_cell(0, 4, sane(f"{explanation['barrier']}: {explanation['message']}"))

    pdf.set_y(277)
    pdf.set_font(font, "I", 6.5)
    pdf.set_text_color(90, 90, 90)
    pdf.multi_cell(0, 3, sane(report["disclaimer"]))
    return bytes(pdf.output())


# ------------------------------------------------------------------------- Excel
def to_xlsx(report: Dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Barriers"
    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="305496")

    cols = ["barrier", "material", "suggested_mm", "B_required", "B_achieved",
            "B_surrogate", "surrogate_CI95", "tier",
            "dose_mSv_wk", "limit_mSv_wk", "verdict", "margin", "engine", "note"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.font = hdr
        cell.fill = fill
    for i, row in enumerate(report["rows"], 2):
        for j, c in enumerate(cols, 1):
            ws.cell(i, j, row[c])
    for col in "ABEFGHIJK":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["K"].width = 40

    meta = wb.create_sheet("Inputs")
    meta.cell(1, 1, report["title"]).font = Font(bold=True, size=13)
    meta.cell(2, 1, report["mode"])
    meta.cell(3, 1, f"Generated {report['timestamp']}")
    r = 5
    for k, v in report["inputs"].items():
        meta.cell(r, 1, k).font = Font(bold=True)
        meta.cell(r, 2, v)
        r += 1
    meta.cell(r + 1, 1, "Disclaimer").font = Font(bold=True)
    meta.cell(r + 2, 1, report["disclaimer"])
    meta.column_dimensions["A"].width = 32
    meta.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -------------------------------------------------------------------------- HTML
def to_html(report: Dict) -> bytes:
    img_b64 = base64.b64encode(report["diagram_png"]).decode() if report.get("diagram_png") else ""
    rows_html = ""
    for row in report["rows"]:
        v = row["verdict"]
        color = {"PASS": "#2e7d32", "FAIL": "#c62828"}.get(v, "#9e9e9e")
        rows_html += (
            f"<tr><td>{row['barrier']}</td><td>{row['material']}</td>"
            f"<td>{row['suggested_mm']}</td><td>{row['B_achieved']}</td>"
            f"<td>{row['B_surrogate']} <span style='color:#666;font-size:11px'>"
            f"{row['surrogate_CI95']}</span></td>"
            f"<td>{row['dose_mSv_wk']}</td><td>{row['limit_mSv_wk']}</td>"
            f"<td style='color:{color};font-weight:bold'>{v}</td>"
            f"<td>{row['margin']}</td><td>{row['tier']}</td></tr>"
        )
    inputs_html = "".join(
        f"<tr><td style='font-weight:bold'>{k}</td><td>{v}</td></tr>"
        for k, v in report["inputs"].items()
    )
    html = f"""<!doctype html><html><head><meta charset="utf-8">
<title>{report['title']}</title>
<style>
 body{{font-family:Arial,Helvetica,sans-serif;margin:32px;color:#222}}
 h1{{font-size:20px}} .sub{{color:#666;font-size:13px}}
 table{{border-collapse:collapse;margin:12px 0}}
 td,th{{border:1px solid #ccc;padding:5px 9px;font-size:13px}}
 th{{background:#305496;color:#fff}}
 .note{{color:#666;font-size:11px;font-style:italic;max-width:720px}}
</style></head><body>
<h1>{report['title']}</h1>
<div class="sub">{report['mode']} &nbsp;|&nbsp; generated {report['timestamp']}</div>
<h3>Inputs</h3><table>{inputs_html}</table>
<img src="data:image/png;base64,{img_b64}" width="520"/>
<h3>Per-barrier results</h3>
<table><tr><th>Barrier</th><th>Material</th><th>Suggested mm</th><th>Analytical B</th>
<th>Surrogate B (95% CI)</th><th>Dose mSv/wk</th>
<th>Limit</th><th>Verdict</th><th>Margin</th><th>Tier</th></tr>{rows_html}</table>
<p class="note">{report['disclaimer']}</p>
{'<p class="note">Notes: ' + report['notes'] + '</p>' if report.get('notes') else ''}
</body></html>"""
    return html.encode("utf-8")


def export(report: Dict, fmt: str) -> tuple:
    """Return (bytes, mime, extension) for the chosen format."""
    fmt = fmt.lower()
    if fmt == "pdf":
        return to_pdf(report), "application/pdf", "pdf"
    if fmt == "excel":
        return (to_xlsx(report),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx")
    return to_html(report), "text/html", "html"
